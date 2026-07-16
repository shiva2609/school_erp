"""
Export utilities — in-memory file generation.

Files are built entirely in RAM and returned as raw bytes / BytesIO so the
caller can stream them directly in an HTTP response without touching the
local filesystem (required for ephemeral cloud environments like DO App
Platform where /media/ is not reliably served).
"""
import html
import io
import os
import uuid

import openpyxl
from openpyxl.styles import Font, PatternFill

from common.pdf_render import html_to_pdf_bytes


# ─── In-memory generators (primary API) ──────────────────────────────────────


def generate_csv_bytes(report_type: str, headers: list, data_rows: list):
    import csv
    import io
    import uuid
    
    string_buffer = io.StringIO()
    writer = csv.writer(string_buffer)
    writer.writerow(headers)
    for row in data_rows:
        writer.writerow(row)
        
    bytes_buffer = io.BytesIO(string_buffer.getvalue().encode('utf-8'))
    file_name = f"{report_type}_{uuid.uuid4().hex[:8]}.csv"
    return bytes_buffer, file_name, "text/csv"

def generate_excel_bytes(report_type: str, headers: list, data_rows: list):
    """
    Build an Excel workbook in memory.
    Returns (BytesIO, filename, content_type).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = str(report_type).title().replace('_', ' ')[:31]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill

    for row_num, row_data in enumerate(data_rows, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=cell_value)

    # Auto-width columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[column].width = max_length + 2

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    file_name = f"{report_type}_{uuid.uuid4().hex[:8]}.xlsx"
    content_type = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    return buffer, file_name, content_type


def generate_pdf_bytes(report_type: str, headers: list, data_rows: list):
    """
    Render tabular report data to PDF bytes via WeasyPrint.
    Returns (bytes, filename, content_type).
    """
    title = str(report_type).replace('_', ' ')
    esc = html.escape
    th = ''.join(f'<th>{esc(str(h))}</th>' for h in headers)
    body_rows = [
        '<tr>' + ''.join(f'<td>{esc(str(v))}</td>' for v in row) + '</tr>'
        for row in data_rows
    ]
    html_string = f"""<!DOCTYPE html>
<html lang="en">
<head><meta charset="utf-8">
<style>
body {{ font-family: DejaVu Sans, Helvetica, Arial, sans-serif; font-size: 10pt; margin: 24px; }}
h1 {{ font-size: 14pt; color: #1e1b4b; }}
table {{ border-collapse: collapse; width: 100%; margin-top: 12px; }}
th, td {{ border: 1px solid #cbd5e1; padding: 6px 8px; text-align: left; vertical-align: top; }}
th {{ background: #4f46e5; color: #fff; }}
tr:nth-child(even) td {{ background: #f8fafc; }}
</style>
</head>
<body>
<h1>{esc(title)}</h1>
<table>
<thead><tr>{th}</tr></thead>
<tbody>{''.join(body_rows)}</tbody>
</table>
</body>
</html>"""

    pdf = html_to_pdf_bytes(html_string)
    file_name = f"{report_type}_{uuid.uuid4().hex[:8]}.pdf"
    return pdf, file_name, "application/pdf"


# ─── Legacy disk-write functions (kept for backward compat) ──────────────────

def generate_excel_file(report_type: str, headers: list, data_rows: list) -> str:
    """
    LEGACY — saves to disk and returns a '/media/exports/…' URL string.
    Prefer generate_excel_bytes() for new code.
    """
    from django.conf import settings

    buffer, file_name, _ = generate_excel_bytes(report_type, headers, data_rows)
    reports_dir = os.path.join(settings.BASE_DIR, 'media', 'exports')
    os.makedirs(reports_dir, exist_ok=True)
    file_path = os.path.join(reports_dir, file_name)
    with open(file_path, 'wb') as fh:
        fh.write(buffer.read())
    return f"/media/exports/{file_name}"


def generate_pdf_file(report_type: str, headers: list, data_rows: list) -> str:
    """
    LEGACY — saves to disk and returns a '/media/exports/…' URL string.
    Prefer generate_pdf_bytes() for new code.
    """
    from django.conf import settings

    pdf_bytes, file_name, _ = generate_pdf_bytes(report_type, headers, data_rows)
    reports_dir = os.path.join(settings.BASE_DIR, 'media', 'exports')
    os.makedirs(reports_dir, exist_ok=True)
    file_path = os.path.join(reports_dir, file_name)
    with open(file_path, 'wb') as fh:
        fh.write(pdf_bytes)
    return f"/media/exports/{file_name}"
