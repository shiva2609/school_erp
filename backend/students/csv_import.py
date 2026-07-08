"""
Student CSV import (onboarding from another SIS).

What each row can create / update
---------------------------------
- **Student**: demographics, address, emergency contact, previous school, parents (inline fields),
  parent **User** links via ``link_parent_accounts_to_student``.
- **Class**: ``grade`` + ``section`` → ``ClassSection`` for the job’s academic year.
- **Admission number**: New students always receive ``admission_number`` from
  ``Student.generate_admission_number`` (tenant ``admission_no_format`` / branch / year).
  The CSV value is stored in ``legacy_admission_number`` for traceability and duplicate matching.
- **Fees (optional columns)**: If ``total_fee`` / ``total amount`` is present, one **annual**
  ``FeeInvoice`` is created with optional ``fee_paid``, ``concession``, ``fee_due_date``,
  and a **Payment** when ``fee_paid`` > 0.
- **Old-year dues**: ``past_due_amount`` + optional ``past_due_year`` creates a ``FeeCarryForward``
  (not a full historical invoice per line item).

Not imported (would need richer templates / multiple rows per student)
---------------------------------------------------------------------
- Per-component fee structures, multiple invoices per year, transport, discounts by head,
  full payment history lines, document files.
"""
import csv
import io
import re
import datetime
from datetime import datetime as dt
from decimal import Decimal, InvalidOperation
from datetime import date
from openpyxl import load_workbook
from django.conf import settings
from django.db import transaction
from django.db.models import Q
from rest_framework.response import Response

from django.core.exceptions import ValidationError as DjangoValidationError

from accounts.permissions import BRANCH_SCOPED_ROLES, normalize_role
from tenants.models import AcademicYear, Branch
from students.models import ClassSection, Student, GRADE_CHOICES, CsvImportJob
from fees.models import FeeInvoice, Payment, FeeCarryForward, DocumentSequence
from .services import create_student_fees, link_parent_accounts_to_student


def _scalar_form_value(val):
    """Normalize multipart / DRF request values (avoid list duplicates, whitespace)."""
    if val is None or val in ('undefined', ''):
        return None
    if isinstance(val, (list, tuple)):
        val = val[0] if val else None
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def handle_csv_import(request):
    """
    Validates CSV file and creates a background CsvImportJob.
    """
    try:
        user = request.user
        branch_id = _scalar_form_value(request.data.get('branch_id'))
        academic_year_id = _scalar_form_value(request.data.get('academic_year_id'))
        file_obj = request.FILES.get('file') or request.data.get('file')
        if not file_obj:
            return Response(
                {
                    'success': False,
                    'detail': 'No file was received. If this keeps happening, try another browser or contact support.',
                },
                status=400,
            )

        file_name = (getattr(file_obj, 'name', None) or '').lower()
        looks_csv = file_name.endswith('.csv')
        looks_xlsx = file_name.endswith('.xlsx')
        if not (looks_csv or looks_xlsx):
            try:
                head = file_obj.read(4)
                if hasattr(file_obj, 'seek'):
                    file_obj.seek(0)
                if len(head) >= 2 and head[:2] == b'PK':
                    looks_xlsx = True
            except Exception:
                if hasattr(file_obj, 'seek'):
                    file_obj.seek(0)

        if not (looks_csv or looks_xlsx):
            return Response({'success': False, 'detail': 'Please upload a valid CSV or XLSX file.'}, status=400)

        max_bytes = getattr(settings, 'STUDENT_CSV_IMPORT_MAX_BYTES', 5 * 1024 * 1024)
        if getattr(file_obj, 'size', 0) and file_obj.size > max_bytes:
            return Response(
                {
                    'success': False,
                    'detail': f'Import file too large. Maximum size is {max_bytes // (1024 * 1024)} MB.',
                },
                status=400,
            )

        role = normalize_role(user.role)

        try:
            if role == 'OWNER':
                if not branch_id:
                    return Response({'success': False, 'detail': 'Owner must provide a branch_id.'}, status=400)
                branch = Branch.objects.get(id=branch_id)
                tenant = branch.tenant
            else:
                if not user.tenant:
                    return Response(
                        {
                            'success': False,
                            'detail': 'Your account is not linked to a school. Contact an administrator.',
                        },
                        status=400,
                    )

                branch = None
                # Branch-scoped roles (accountant, teacher, etc.): always use the assigned branch.
                # Client branch_id comes from localStorage and is often stale → wrong id caused
                # "Invalid branch or academic year" even when the UI looks correct.
                if role in BRANCH_SCOPED_ROLES and getattr(user, 'branch_id', None):
                    branch = user.branch
                elif branch_id:
                    branch = Branch.objects.get(id=branch_id, tenant=user.tenant)
                else:
                    branch = user.branch

                if not branch:
                    return Response(
                        {
                            'success': False,
                            'detail': 'No branch associated with your account. Pick a branch in the header or ask an admin to assign you to a branch.',
                        },
                        status=400,
                    )
                tenant = branch.tenant

            if academic_year_id:
                ay = AcademicYear.objects.get(id=academic_year_id, tenant=tenant)
            else:
                ay = AcademicYear.objects.filter(tenant=tenant, is_active=True).first()
                if not ay:
                    return Response({'success': False, 'detail': 'No active academic year found. Please select one.'}, status=400)
        except Branch.DoesNotExist:
            return Response(
                {
                    'success': False,
                    'detail': 'That branch was not found for your school. Clear site data or pick another branch in the header, then try again.',
                },
                status=400,
            )
        except AcademicYear.DoesNotExist:
            return Response(
                {
                    'success': False,
                    'detail': 'That academic year was not found for this school. Refresh the page and choose the year again.',
                },
                status=400,
            )
        except (ValueError, DjangoValidationError):
            return Response(
                {'success': False, 'detail': 'Invalid branch or academic year id. Refresh and try again.'},
                status=400,
            )

        update_student_details = str(request.data.get('update_student_details', '')).lower() == 'true'
        update_fee_details = str(request.data.get('update_fee_details', '')).lower() == 'true'

        # Create the background job
        job = CsvImportJob.objects.create(
            tenant=tenant,
            branch=branch,
            academic_year=ay,
            file=file_obj,
            created_by=user,
            status='PENDING',
            update_student_details=update_student_details,
            update_fee_details=update_fee_details
        )

        # Trigger Celery Task
        from .tasks import process_student_csv_import
        process_student_csv_import.delay(job.id)

        return Response({
            'success': True,
            'message': 'Import started in the background.',
            'job_id': job.id
        })

    except Exception as e:
        import logging
        logging.getLogger(__name__).exception('CSV import request failed')
        return Response({
            'success': False,
            'detail': 'An unexpected server error occurred.',
        }, status=500)


def process_rows(job, rows):
    """Actual logic to process parsed rows asynchronously."""
    if not rows:
        job.status = 'FAILED'
        job.error_log = ['Import file is empty.']
        job.save(update_fields=['status', 'error_log'])
        return

    job.total_rows = len(rows)
    job.status = 'PROCESSING'
    job.save(update_fields=['total_rows', 'status'])

    errors = []
    success_count = 0
    skipped_duplicates = 0
    updated_count = 0
    processed_rows = 0

    tenant = job.tenant
    branch = job.branch
    ay = job.academic_year
    user = job.created_by

    # ── helpers ──────────────────────────────────────────────────────────
    def parse_date(date_str):
        if not date_str:
            return None
        date_str = date_str.strip()
        for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y', '%Y/%m/%d', '%d.%m.%Y']:
            try:
                return dt.strptime(date_str, fmt).date()
            except ValueError:
                continue
        return None

    def get_val(row, *keys):
        """Flexible column lookup: exact, underscore→space, partial match."""
        for k in keys:
            if k in row and row[k]:
                return row[k]
            k_space = k.replace('_', ' ')
            if k_space in row and row[k_space]:
                return row[k_space]
            for rk in row.keys():
                if k == rk or k_space == rk:
                    continue
                if k in rk and row[rk]:
                    return row[rk]
                if k_space in rk and row[rk]:
                    return row[rk]
        return ''

    def safe_phone(val, max_len=15):
        if not val:
            return None
        cleaned = re.sub(r'[\s\-\(\)]', '', str(val))
        return cleaned[:max_len] if cleaned else None

    def safe_str(val, max_len=None):
        if not val:
            return None
        s = str(val).strip()
        if max_len:
            s = s[:max_len]
        return s or None

    def parse_decimal(val_str):
        if not val_str:
            return Decimal('0')
        val_str = str(val_str).strip()
        if '(' in val_str:
            val_str = val_str.split('(')[0]
        cleaned = re.sub(r'[^\d\.\-]', '', val_str)
        if not cleaned:
            return Decimal('0')
        try:
            return Decimal(cleaned)
        except InvalidOperation:
            return Decimal('0')

    # Chunk processing: batch size of 50
    CHUNK_SIZE = 50
    
    for i in range(0, len(rows), CHUNK_SIZE):
        chunk = rows[i:i + CHUNK_SIZE]
        
        for row_idx_offset, raw_row in enumerate(chunk):
            row_idx = i + row_idx_offset + 2 # +2 for header and 1-based index
            row_label = f"Row {row_idx}"
            
            try:
                with transaction.atomic():
                    row = {}
                    for k, v in raw_row.items():
                        key = k.strip().lower() if isinstance(k, str) and k else 'col'
                        if isinstance(v, list):
                            val = ' '.join(str(x) for x in v if x).strip()
                        elif isinstance(v, str):
                            val = v.strip()
                        else:
                            val = str(v).strip() if v else ''
                        row[key] = val

                    if not any(row.values()):
                        processed_rows += 1
                        continue

                    first_name = get_val(row, 'first name', 'first_name', 'student name', 'name').strip()
                    if not first_name:
                        raise ValueError("Student name is required.")
                    last_name  = get_val(row, 'last name', 'last_name').strip()
                    if not last_name and ' ' in first_name:
                        parts = first_name.rsplit(' ', 1)
                        first_name = parts[0].strip()
                        last_name  = parts[1].strip()

                    row_label = f"Row {row_idx} ({first_name} {last_name})".strip()

                    dob_raw   = get_val(row, 'date of birth', 'date_of_birth', 'dob').strip()
                    gender    = get_val(row, 'gender').strip().upper()
                    grade_str = get_val(row, 'class', 'grade', 'class name').strip()

                    section_raw = get_val(row, 'section').strip()
                    section = section_raw.split()[0] if section_raw else 'A'
                    section = section[:50]

                    csv_admission = get_val(row, 'admission number', 'admission_number', 'admission no', 'old admission', 'legacy admission').strip()

                    if gender not in ('MALE', 'FEMALE', 'OTHER'):
                        gender = 'OTHER'

                    parsed_dob = parse_date(dob_raw) or date(2000, 1, 1)

                    grade = None
                    if grade_str:
                        g_clean = grade_str.upper().strip().replace('GRADE', '').replace('CLASS', '').strip()
                        roman_map = {'I': '1', 'II': '2', 'III': '3', 'IV': '4', 'V': '5', 'VI': '6', 'VII': '7', 'VIII': '8', 'IX': '9', 'X': '10', 'XI': '11', 'XII': '12'}
                        if g_clean in roman_map: g_clean = roman_map[g_clean]
                        for k, v in GRADE_CHOICES:
                            if g_clean == k or g_clean == v.upper().replace('GRADE ', ''):
                                grade = k
                                break
                        if not grade and grade_str.upper().strip() in dict(GRADE_CHOICES):
                            grade = grade_str.upper().strip()
                        if not grade: grade = grade_str[:50]

                    cs = None
                    if grade:
                        cs, _ = ClassSection.objects.get_or_create(
                            tenant=tenant, branch=branch, academic_year=ay,
                            grade=grade, section=section.upper(),
                        )

                    existing_student = None
                    if csv_admission:
                        existing_student = Student.objects.filter(
                            branch=branch, academic_year=ay,
                        ).filter(
                            Q(admission_number__iexact=csv_admission)
                            | Q(legacy_admission_number__iexact=csv_admission),
                        ).first()
                    if not existing_student and cs:
                        existing_student = Student.objects.filter(
                            branch=branch, academic_year=ay,
                            first_name__iexact=first_name, last_name__iexact=last_name,
                            date_of_birth=parsed_dob, class_section=cs,
                        ).first()

                    blood_group_raw = row.get('blood group', row.get('blood_group', '')).upper().strip()
                    blood_group = blood_group_raw if blood_group_raw in ['A+', 'A-', 'B+', 'B-', 'AB+', 'AB-', 'O+', 'O-', 'UNKNOWN', ''] else 'UNKNOWN'
                    caste_raw = row.get('caste category', row.get('caste_category', '')).upper().strip()
                    caste_category = caste_raw if caste_raw in ['GEN', 'OBC', 'SC', 'ST', 'EWS', 'OTHER', ''] else None

                    roll_raw = get_val(row, 'roll number', 'roll_number', 'roll no').strip()
                    roll_num = int(roll_raw) if roll_raw.isdigit() else None
                    
                    religion = safe_str(row.get('religion'), 100)
                    aadhar_number = safe_str(row.get('aadhar number', row.get('aadhar_number')), 12)
                    mother_tongue = safe_str(row.get('mother tongue', row.get('mother_tongue')), 50)
                    nationality = safe_str(row.get('nationality'), 50) or 'Indian'
                    father_name = safe_str(get_val(row, 'father name', 'father_name', 'parent name'), 200)
                    father_phone = safe_phone(get_val(row, 'father mobile', 'father_phone', 'parent mobile', 'father mobile'))
                    father_email = safe_str(row.get('father email', row.get('father_email')), 254)
                    father_qualification = safe_str(row.get('father_qualification'), 100)
                    father_occupation = safe_str(row.get('father_occupation'), 100)
                    father_aadhaar = safe_str(row.get('father_aadhaar'), 12)
                    mother_name = safe_str(get_val(row, 'mother name', 'mother_name'), 200)
                    mother_phone = safe_phone(get_val(row, 'mother mobile', 'mother_phone', 'mother mobile'))
                    mother_email = safe_str(row.get('mother email', row.get('mother_email')), 254)
                    mother_qualification = safe_str(row.get('mother_qualification'), 100)
                    mother_occupation = safe_str(row.get('mother_occupation'), 100)
                    mother_aadhaar = safe_str(row.get('mother_aadhaar'), 12)
                    guardian_name = safe_str(get_val(row, 'guardian name', 'guardian_name'), 200)
                    guardian_phone = safe_phone(get_val(row, 'guardian mobile', 'guardian_phone', 'guardian mobile'))
                    guardian_relation = safe_str(row.get('guardian_relation'), 100)
                    address_line1 = safe_str(row.get('address', row.get('address_line1')), 255)
                    address_line2 = safe_str(row.get('address_line2'), 255)
                    city = safe_str(row.get('city'), 100)
                    district = safe_str(row.get('district'), 100)
                    state = safe_str(row.get('state'), 100)
                    pincode = safe_str(row.get('pincode'), 6)
                    previous_school_name = safe_str(row.get('previous_school_name'), 200)
                    previous_class = safe_str(row.get('previous_class'), 20)
                    previous_school_ay = safe_str(row.get('previous_school_ay'), 20)
                    emergency_contact_name = safe_str(row.get('emergency_contact_name'), 200)
                    emergency_contact_phone = safe_phone(row.get('emergency_contact_phone'))
                    emergency_contact_relation = safe_str(row.get('emergency_contact_relation'), 100)
                    
                    is_new_student = False
                    is_student_updated = False
                    
                    if existing_student:
                        student = existing_student
                        if job.update_student_details:
                            student.first_name = first_name
                            student.last_name = last_name or ''
                            student.date_of_birth = parsed_dob
                            student.gender = gender
                            if grade:
                                student.grade = grade
                            if cs:
                                student.class_section = cs
                            student.roll_number = roll_num
                            student.blood_group = blood_group or 'UNKNOWN'
                            student.religion = religion
                            student.caste_category = caste_category
                            student.aadhar_number = aadhar_number
                            student.mother_tongue = mother_tongue
                            student.nationality = nationality
                            student.father_name = father_name
                            student.father_phone = father_phone
                            student.father_email = father_email
                            student.father_qualification = father_qualification
                            student.father_occupation = father_occupation
                            student.father_aadhaar = father_aadhaar
                            student.mother_name = mother_name
                            student.mother_phone = mother_phone
                            student.mother_email = mother_email
                            student.mother_qualification = mother_qualification
                            student.mother_occupation = mother_occupation
                            student.mother_aadhaar = mother_aadhaar
                            student.guardian_name = guardian_name
                            student.guardian_phone = guardian_phone
                            student.guardian_relation = guardian_relation
                            student.address_line1 = address_line1
                            student.address_line2 = address_line2
                            student.city = city
                            student.district = district
                            student.state = state
                            student.pincode = pincode
                            student.previous_school_name = previous_school_name
                            student.previous_class = previous_class
                            student.previous_school_ay = previous_school_ay
                            student.emergency_contact_name = emergency_contact_name
                            student.emergency_contact_phone = emergency_contact_phone
                            student.emergency_contact_relation = emergency_contact_relation
                            student.save()
                            
                            father_info = {'phone': student.father_phone, 'email': student.father_email, 'name': student.father_name or ''}
                            mother_info = {'phone': student.mother_phone, 'email': student.mother_email, 'name': student.mother_name or ''}
                            link_parent_accounts_to_student(
                                student, father_info, mother_info, tenant, branch,
                                strict_parent_email=False,
                            )
                            is_student_updated = True
                            updated_count += 1
                        else:
                            if not job.update_fee_details:
                                skipped_duplicates += 1
                    else:
                        is_new_student = True
                        platform_admission = Student.generate_admission_number(branch, ay)
                        while Student.objects.filter(
                            branch=branch, academic_year=ay, admission_number=platform_admission,
                        ).exists():
                            platform_admission = Student.generate_admission_number(branch, ay)
                        legacy_stored = (csv_admission or '')[:64]

                        student = Student.objects.create(
                            tenant=tenant, branch=branch, academic_year=ay, grade=grade, class_section=cs,
                            first_name=first_name, last_name=last_name or '', date_of_birth=parsed_dob,
                            gender=gender, admission_number=platform_admission,
                            legacy_admission_number=legacy_stored,
                            roll_number=roll_num,
                            blood_group=blood_group or 'UNKNOWN', religion=religion, caste_category=caste_category,
                            aadhar_number=aadhar_number, mother_tongue=mother_tongue,
                            nationality=nationality,
                            father_name=father_name,
                            father_phone=father_phone,
                            father_email=father_email, father_qualification=father_qualification,
                            father_occupation=father_occupation, father_aadhaar=father_aadhaar,
                            mother_name=mother_name,
                            mother_phone=mother_phone,
                            mother_email=mother_email, mother_qualification=mother_qualification,
                            mother_occupation=mother_occupation, mother_aadhaar=mother_aadhaar,
                            guardian_name=guardian_name,
                            guardian_phone=guardian_phone,
                            guardian_relation=guardian_relation,
                            address_line1=address_line1, address_line2=address_line2,
                            city=city, district=district,
                            state=state, pincode=pincode,
                            previous_school_name=previous_school_name, previous_class=previous_class,
                            previous_school_ay=previous_school_ay,
                            emergency_contact_name=emergency_contact_name, emergency_contact_phone=emergency_contact_phone,
                            emergency_contact_relation=emergency_contact_relation,
                            created_by=user, status='ACTIVE',
                        )

                        father_info = {'phone': student.father_phone, 'email': student.father_email, 'name': student.father_name or ''}
                        mother_info = {'phone': student.mother_phone, 'email': student.mother_email, 'name': student.mother_name or ''}
                        link_parent_accounts_to_student(
                            student, father_info, mother_info, tenant, branch,
                            strict_parent_email=False,
                        )

                    # Determine whether granular columns are present
                    has_granular = any(k in row for k in ['tuition fee', 'transport fee'])
                    past_due_year_raw = get_val(row, 'past_due_year', 'past due year', 'arrears year').strip()
                    fee_due_date_raw = get_val(row, 'fee_due_date', 'due date', 'fee due date').strip()

                    has_fee_data = False
                    
                    # Initialize all fee fields to Decimal('0')
                    tuition_fee = Decimal('0')
                    tuition_concession = Decimal('0')
                    tuition_collected = Decimal('0')
                    
                    transport_fee = Decimal('0')
                    transport_concession = Decimal('0')
                    transport_collected = Decimal('0')
                    
                    past_due = Decimal('0')
                    past_due_concession = Decimal('0')
                    past_due_collected = Decimal('0')

                    if has_granular:
                        tuition_fee = parse_decimal(get_val(row, 'tuition fee'))
                        transport_fee = parse_decimal(get_val(row, 'transport fee'))
                        past_due = parse_decimal(get_val(row, 'past due'))

                        tuition_collected = parse_decimal(get_val(row, 'tuition collected'))
                        transport_collected = parse_decimal(get_val(row, 'transport collected'))
                        past_due_collected = parse_decimal(get_val(row, 'past due collected'))

                        tuition_concession = parse_decimal(get_val(row, 'tuition concession'))
                        transport_concession = parse_decimal(get_val(row, 'transport concession'))
                        past_due_concession = parse_decimal(get_val(row, 'past due concession'))
                        
                        if tuition_fee > 0 or transport_fee > 0 or past_due > 0:
                            has_fee_data = True
                    else:
                        total_fee_raw = get_val(row, 'total_fee', 'total amount (₹)', 'total fee', 'total amount')
                        if total_fee_raw:
                            tuition_fee = parse_decimal(total_fee_raw)
                            tuition_collected = parse_decimal(get_val(row, 'fee_paid', 'amount paid (₹)', 'fee paid', 'total paid'))
                            tuition_concession = parse_decimal(get_val(row, 'concession_amount', 'concession (₹)', 'concession', 'total concession'))
                            
                            past_due = parse_decimal(get_val(row, 'past_due_amount', 'past due', 'old dues', 'arrears'))
                            past_due_collected = parse_decimal(get_val(row, 'past_due_collected', 'past due collected', 'old dues paid', 'arrears paid'))
                            past_due_concession = parse_decimal(get_val(row, 'past_due_concession', 'past due concession', 'old dues concession', 'arrears concession'))
                            
                            has_fee_data = True

                    if has_fee_data:
                        # 1. Clamp Concessions to never exceed standard fee amount
                        tuition_concession = min(tuition_concession, tuition_fee)
                        transport_concession = min(transport_concession, transport_fee)
                        past_due_concession = min(past_due_concession, past_due)
                        
                        # 2. Calculate Accepted/Net amounts
                        accepted_tuition = tuition_fee - tuition_concession
                        accepted_transport = transport_fee - transport_concession
                        accepted_past_due = past_due - past_due_concession
                        
                        # 3. Clamp Collections to never exceed accepted net amounts
                        tuition_collected = min(tuition_collected, accepted_tuition)
                        transport_collected = min(transport_collected, accepted_transport)
                        past_due_collected = min(past_due_collected, accepted_past_due)

                        from fees.models import FeeCategory as FC, FeeInvoiceItem, StudentFeeItem, PaymentAllocation
                        
                        # Get/Create Tuition Category
                        tuition_cat, _ = FC.objects.get_or_create(
                            branch=branch,
                            code='TUITION',
                            defaults={
                                'tenant': tenant,
                                'name': 'Tuition Fee',
                                'description': 'Academic tuition fee',
                                'is_active': True,
                                'order': 1,
                            }
                        )
                        
                        # Get/Create Transport Category
                        transport_cat, _ = FC.objects.get_or_create(
                            branch=branch,
                            code='TRANSPORT',
                            defaults={
                                'tenant': tenant,
                                'name': 'Transport Fee',
                                'description': 'Monthly school transport fee',
                                'is_active': True,
                                'order': 99,
                            }
                        )

                        tuition_invoice = None
                        due_date = parse_date(fee_due_date_raw) or date.today()
                        
                        # We create a Tuition Invoice if tuition_fee > 0 or if we need a placeholder anchor for past dues
                        need_placeholder = (tuition_fee == 0 and transport_fee == 0 and past_due > 0 and past_due_collected > 0)
                        
                        if tuition_fee > 0 or need_placeholder:
                            # Check if annual tuition invoice already exists
                            tuition_invoice = FeeInvoice.objects.filter(
                                student=student, academic_year=ay, month="ANNUAL"
                            ).exclude(invoice_number__startswith='TRN-').exclude(
                                invoice_number__startswith='ADM-'
                            ).exclude(
                                invoice_number__startswith='FDP-'
                            ).exclude(
                                invoice_number__startswith='SPF-'
                            ).first()
                            
                            if not tuition_invoice or job.update_fee_details:
                                tuition_net = accepted_tuition
                                tuition_status = 'PAID' if tuition_net <= 0 or tuition_collected >= tuition_net else ('PARTIALLY_PAID' if tuition_collected > 0 else 'SENT')
                                if tuition_status not in ('PAID', 'CANCELLED', 'WAIVED') and due_date < date.today():
                                    tuition_status = 'OVERDUE'
                                    
                                if not tuition_invoice:
                                    tuition_inv_number = DocumentSequence.get_next_sequence(branch, 'INVOICE', f"INV-{ay.start_date.year:04d}")
                                    tuition_invoice = FeeInvoice.objects.create(
                                        tenant=tenant, branch=branch, academic_year=ay, student=student,
                                        invoice_number=tuition_inv_number, month="ANNUAL",
                                        gross_amount=tuition_fee, concession_amount=tuition_concession, net_amount=tuition_net,
                                        paid_amount=tuition_collected, outstanding_amount=max(Decimal('0'), tuition_net - tuition_collected),
                                        due_date=due_date, status=tuition_status, generated_by='MANUAL', created_by=user,
                                    )
                                else:
                                    tuition_invoice.gross_amount = tuition_fee
                                    tuition_invoice.concession_amount = tuition_concession
                                    tuition_invoice.net_amount = tuition_net
                                    tuition_invoice.paid_amount = tuition_collected
                                    tuition_invoice.outstanding_amount = max(Decimal('0'), tuition_net - tuition_collected)
                                    tuition_invoice.due_date = due_date
                                    tuition_invoice.status = tuition_status
                                    tuition_invoice.save()
                                
                                # Create or Update FeeInvoiceItem
                                item = FeeInvoiceItem.objects.filter(invoice=tuition_invoice, category=tuition_cat).first()
                                if item:
                                    item.original_amount = tuition_fee
                                    item.concession = tuition_concession
                                    item.final_amount = accepted_tuition
                                    item.save()
                                else:
                                    FeeInvoiceItem.objects.create(
                                        invoice=tuition_invoice,
                                        category=tuition_cat,
                                        original_amount=tuition_fee,
                                        concession=tuition_concession,
                                        final_amount=accepted_tuition,
                                        description='Annual tuition fee',
                                    )
                                
                                # Create or Update StudentFeeItem
                                sf_item = StudentFeeItem.objects.filter(student=student, academic_year=ay, category=tuition_cat).first()
                                if sf_item:
                                    sf_item.amount = accepted_tuition
                                    sf_item.save()
                                else:
                                    StudentFeeItem.objects.create(
                                        student=student,
                                        academic_year=ay,
                                        category=tuition_cat,
                                        amount=accepted_tuition,
                                        is_locked=True
                                    )
                                
                                # Create or Update Payment if tuition_collected > 0
                                if tuition_collected > 0:
                                    payment = Payment.objects.filter(tenant=tenant, invoice=tuition_invoice).first()
                                    if payment:
                                        payment.amount = tuition_collected
                                        payment.payment_date = date.today()
                                        payment.save()
                                        allocation = PaymentAllocation.objects.filter(payment=payment, invoice=tuition_invoice).first()
                                        if allocation:
                                            allocation.allocated_amount = tuition_collected
                                            allocation.save()
                                    else:
                                        receipt_number = DocumentSequence.get_next_sequence(branch, 'RECEIPT', f"RCP-{branch.branch_code.upper().replace(' ', '')}-{ay.start_date.year:04d}")
                                        payment = Payment.objects.create(
                                            tenant=tenant, branch=branch, invoice=tuition_invoice, student=student,
                                            amount=tuition_collected, payment_mode='CASH', payment_date=date.today(),
                                            status='COMPLETED', collected_by=user, receipt_number=receipt_number,
                                        )
                                        PaymentAllocation.objects.create(
                                            payment=payment,
                                            invoice=tuition_invoice,
                                            allocated_amount=tuition_collected,
                                            allocation_type='CURRENT_YEAR',
                                        )

                        transport_invoice = None
                        if transport_fee > 0:
                            # Check if annual transport invoice already exists
                            transport_invoice = FeeInvoice.objects.filter(
                                student=student, academic_year=ay, month="ANNUAL", invoice_number__startswith='TRN-'
                            ).first()
                            
                            if not transport_invoice or job.update_fee_details:
                                transport_net = accepted_transport
                                transport_status = 'PAID' if transport_net <= 0 or transport_collected >= transport_net else ('PARTIALLY_PAID' if transport_collected > 0 else 'SENT')
                                if transport_status not in ('PAID', 'CANCELLED', 'WAIVED') and due_date < date.today():
                                    transport_status = 'OVERDUE'
                                    
                                if not transport_invoice:
                                    transport_inv_number = DocumentSequence.get_next_sequence(branch, 'INVOICE', f"TRN-{student.branch.branch_code}-{ay.start_date.year:04d}")
                                    transport_invoice = FeeInvoice.objects.create(
                                        tenant=tenant, branch=branch, academic_year=ay, student=student,
                                        invoice_number=transport_inv_number, month="ANNUAL",
                                        gross_amount=transport_fee, concession_amount=transport_concession, net_amount=transport_net,
                                        paid_amount=transport_collected, outstanding_amount=max(Decimal('0'), transport_net - transport_collected),
                                        due_date=due_date, status=transport_status, generated_by='MANUAL', created_by=user,
                                    )
                                else:
                                    transport_invoice.gross_amount = transport_fee
                                    transport_invoice.concession_amount = transport_concession
                                    transport_invoice.net_amount = transport_net
                                    transport_invoice.paid_amount = transport_collected
                                    transport_invoice.outstanding_amount = max(Decimal('0'), transport_net - transport_collected)
                                    transport_invoice.due_date = due_date
                                    transport_invoice.status = transport_status
                                    transport_invoice.save()
                                
                                # Create or Update FeeInvoiceItem
                                item = FeeInvoiceItem.objects.filter(invoice=transport_invoice, category=transport_cat).first()
                                if item:
                                    item.original_amount = transport_fee
                                    item.concession = transport_concession
                                    item.final_amount = accepted_transport
                                    item.save()
                                else:
                                    FeeInvoiceItem.objects.create(
                                        invoice=transport_invoice,
                                        category=transport_cat,
                                        original_amount=transport_fee,
                                        concession=transport_concession,
                                        final_amount=accepted_transport,
                                        description='Annual transport fee',
                                    )
                                
                                # Create or Update StudentFeeItem
                                sf_item = StudentFeeItem.objects.filter(student=student, academic_year=ay, category=transport_cat).first()
                                if sf_item:
                                    sf_item.amount = accepted_transport
                                    sf_item.save()
                                else:
                                    StudentFeeItem.objects.create(
                                        student=student,
                                        academic_year=ay,
                                        category=transport_cat,
                                        amount=accepted_transport,
                                        is_locked=True
                                    )
                                
                                # Create or Update Payment if transport_collected > 0
                                if transport_collected > 0:
                                    payment = Payment.objects.filter(tenant=tenant, invoice=transport_invoice).first()
                                    if payment:
                                        payment.amount = transport_collected
                                        payment.payment_date = date.today()
                                        payment.save()
                                        allocation = PaymentAllocation.objects.filter(payment=payment, invoice=transport_invoice).first()
                                        if allocation:
                                            allocation.allocated_amount = transport_collected
                                            allocation.save()
                                    else:
                                        receipt_number = DocumentSequence.get_next_sequence(branch, 'RECEIPT', f"RCP-{branch.branch_code.upper().replace(' ', '')}-{ay.start_date.year:04d}")
                                        payment = Payment.objects.create(
                                            tenant=tenant, branch=branch, invoice=transport_invoice, student=student,
                                            amount=transport_collected, payment_mode='CASH', payment_date=date.today(),
                                            status='COMPLETED', collected_by=user, receipt_number=receipt_number,
                                        )
                                        PaymentAllocation.objects.create(
                                            payment=payment,
                                            invoice=transport_invoice,
                                            allocated_amount=transport_collected,
                                            allocation_type='CURRENT_YEAR',
                                        )

                        if past_due > 0:
                            legacy_ay_name = past_due_year_raw or "Legacy-Dues"
                            target_year    = ay.start_date.year - 1
                            legacy_ay, _   = AcademicYear.objects.get_or_create(
                                tenant=tenant, name=legacy_ay_name,
                                defaults={'start_date': datetime.date(target_year, 4, 1), 'end_date': datetime.date(target_year + 1, 3, 31), 'is_active': False, 'status': 'CLOSED'}
                            )

                            # Check if carry forward already exists
                            cf = FeeCarryForward.objects.filter(
                                student=student, source_academic_year=legacy_ay, target_academic_year=ay
                            ).first()

                            if not cf or job.update_fee_details:
                                remaining_cf = past_due - past_due_collected - past_due_concession
                                if remaining_cf <= 0:
                                    cf_status = 'PAID' if past_due_collected > 0 else ('WRITTEN_OFF' if past_due_concession > 0 else 'PAID')
                                elif past_due_collected > 0:
                                    cf_status = 'PARTIALLY_PAID'
                                else:
                                    cf_status = 'PENDING'

                                if not cf:
                                    cf = FeeCarryForward.objects.create(
                                        tenant=tenant, branch=branch, student=student, source_academic_year=legacy_ay, target_academic_year=ay,
                                        total_fee_amount=past_due, total_paid_amount=Decimal('0.00'), carry_forward_amount=past_due,
                                        paid_amount=past_due_collected, written_off_amount=past_due_concession, status=cf_status, created_by=user,
                                    )
                                else:
                                    cf.total_fee_amount = past_due
                                    cf.carry_forward_amount = past_due
                                    cf.paid_amount = past_due_collected
                                    cf.written_off_amount = past_due_concession
                                    cf.status = cf_status
                                    cf.save()

                                if past_due_collected > 0:
                                    # Anchor payment to Tuition Invoice or Transport Invoice
                                    anchor_invoice = tuition_invoice or transport_invoice
                                    if anchor_invoice:
                                        payment_cf = Payment.objects.filter(tenant=tenant, student=student, allocations__carry_forward=cf).first()
                                        if payment_cf:
                                            payment_cf.amount = past_due_collected
                                            payment_cf.payment_date = date.today()
                                            payment_cf.save()
                                            allocation = PaymentAllocation.objects.filter(payment=payment_cf, carry_forward=cf).first()
                                            if allocation:
                                                allocation.allocated_amount = past_due_collected
                                                allocation.save()
                                        else:
                                            receipt_number = DocumentSequence.get_next_sequence(branch, 'RECEIPT', f"RCP-{branch.branch_code.upper().replace(' ', '')}-{ay.start_date.year:04d}")
                                            payment_cf = Payment.objects.create(
                                                tenant=tenant, branch=branch, invoice=anchor_invoice, student=student,
                                                amount=past_due_collected, payment_mode='CASH', payment_date=date.today(),
                                                status='COMPLETED', collected_by=user, receipt_number=receipt_number,
                                            )
                                            PaymentAllocation.objects.create(
                                                payment=payment_cf,
                                                carry_forward=cf,
                                                allocated_amount=past_due_collected,
                                                allocation_type='PREVIOUS_YEAR_DUES',
                                            )
                    else:
                        if is_new_student:
                            create_student_fees(student, None, None, 'Auto-generated on CSV Import', user)

                    success_count += 1
                    
            except Exception as row_error:
                import traceback
                traceback.print_exc()
                errors.append(f"{row_label}: {str(row_error)}")
            
            processed_rows += 1
            
        # Update job progress after every chunk
        job.processed_rows = processed_rows
        job.success_count = success_count
        job.skipped_duplicates = skipped_duplicates
        job.updated_count = updated_count
        job.error_log = errors
        job.save(update_fields=['processed_rows', 'success_count', 'skipped_duplicates', 'updated_count', 'error_log'])

    # Finalize job
    job.status = 'COMPLETED'
    job.save(update_fields=['status'])


def process_csv_file(job, decoded_file):
    """Parse CSV content and process rows."""
    io_string = io.StringIO(decoded_file)
    reader = list(csv.DictReader(io_string))

    if not reader:
        job.status = 'FAILED'
        job.error_log = ['CSV file is empty.']
        job.save(update_fields=['status', 'error_log'])
        return

    # Normalize headers
    fieldnames = reader[0].keys() if reader else []
    normalized_headers = [h.strip().lower() if h else f'col_{i}' for i, h in enumerate(fieldnames)]

    # Recreate reader with normalized headers
    io_string.seek(0)
    reader_obj = csv.DictReader(io_string, fieldnames=normalized_headers)
    next(reader_obj)  # skip header row
    rows = list(reader_obj)
    process_rows(job, rows)


def process_xlsx_file(job, raw_bytes):
    """Parse XLSX content and process rows."""
    workbook = load_workbook(filename=io.BytesIO(raw_bytes), read_only=True, data_only=True)
    sheet = workbook.active
    row_iter = sheet.iter_rows(values_only=True)

    try:
        header_row = next(row_iter)
    except StopIteration:
        workbook.close()
        job.status = 'FAILED'
        job.error_log = ['XLSX file is empty.']
        job.save(update_fields=['status', 'error_log'])
        return

    headers = []
    for idx, cell in enumerate(header_row):
        if cell is None:
            headers.append(f'col_{idx}')
        else:
            headers.append(str(cell).strip().lower() or f'col_{idx}')

    rows = []
    for values in row_iter:
        row = {}
        for idx, header in enumerate(headers):
            val = values[idx] if idx < len(values) else None
            if val is None:
                str_val = ''
            elif hasattr(val, 'strftime'):
                str_val = val.strftime('%Y-%m-%d')
            elif isinstance(val, float) and val.is_integer():
                str_val = str(int(val))
            else:
                str_val = str(val).strip()
            row[header] = str_val
        rows.append(row)

    workbook.close()
    process_rows(job, rows)
